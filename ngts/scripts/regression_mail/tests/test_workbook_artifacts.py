from __future__ import annotations

import hashlib
import json
import tempfile
import unittest
from pathlib import Path

import openpyxl

from ngts.scripts.regression_mail.grouping import build_deterministic_report
from ngts.scripts.regression_mail.models import SemanticGroup
from ngts.scripts.regression_mail.workbook import (
    ArtifactWriter,
    FALLBACK_SKIP_REASON,
    REQUIRED_HEADERS,
    WorkbookLoader,
)


VERSION = "202608_RC.17-deadbeef_Internal"


class WorkbookArtifactTest(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(self.temp_dir.cleanup)
        self.path = Path(self.temp_dir.name) / "results.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "community_tests"
        headers = list(REQUIRED_HEADERS) + ["extra"]
        sheet.append(headers)
        sheet.append(
            [
                "s1",
                "k1",
                "tb1",
                "tests/a.py::test_x[a]",
                "fail",
                "failure text",
                "t0",
                "dut",
                "spc",
                "x86",
                "sku",
                "SONiC." + VERSION,
                "tests/a.py::test_x[a]",
                "keep",
            ]
        )
        sheet.append(
            [
                "s2",
                "k2",
                "tb2",
                "tests/b.py::test_skip",
                "skipped",
                "Internal ticket",
                "t1",
                "dut",
                "spc",
                "x86",
                "sku2",
                VERSION,
                "tests/b.py::test_skip",
                "keep2",
            ]
        )
        sheet.append(
            [
                "s3",
                "k3",
                "tb3",
                "tests/c.py::test_other",
                "pass",
                "",
                "t2",
                "dut",
                "spc",
                "x86",
                "sku3",
                "202605_RC.1-other_Internal",
                "tests/c.py::test_other",
                "other-version",
            ]
        )
        workbook.create_sheet("metadata")["A1"] = "preserve"
        workbook.save(self.path)
        workbook.close()

    def test_filters_exact_version_and_preserves_source(self) -> None:
        before = hashlib.sha256(self.path.read_bytes()).hexdigest()
        snapshot = WorkbookLoader().load(self.path, VERSION)
        after = hashlib.sha256(self.path.read_bytes()).hexdigest()

        self.assertEqual(before, after)
        self.assertEqual("community_tests", snapshot.sheet_name)
        self.assertEqual(2, len(snapshot.selected_rows))
        self.assertEqual({"fail": 1, "pass": 0, "skipped": 1}, snapshot.result_counts)
        self.assertEqual([("sku", "t0"), ("sku2", "t1")], snapshot.hardware_pairs)

    def test_attachment_and_atomic_skip_mapping(self) -> None:
        snapshot = WorkbookLoader().load(self.path, VERSION)
        semantic = build_deterministic_report(snapshot)
        skip_group = semantic.skip_groups[0]
        semantic.skip_groups[0] = SemanticGroup(
            group_id=skip_group.group_id,
            member_ids=skip_group.member_ids,
            test_display=skip_group.test_display,
            testbeds=skip_group.testbeds,
            comments="",
            internal_comments="https://redmine.mellanox.com/issues/123",
            redmine_urls=["https://redmine.mellanox.com/issues/123"],
        )
        writer = ArtifactWriter()
        skips_path = writer.write_skips(snapshot, semantic)
        attachment = writer.build_attachment(snapshot, semantic)
        self.addCleanup(writer.cleanup_attachment, attachment)

        self.assertEqual(
            {"https://redmine.mellanox.com/issues/123": "Internal ticket"},
            json.loads(skips_path.read_text(encoding="utf-8")),
        )
        enriched = openpyxl.load_workbook(attachment, read_only=True)
        try:
            self.assertIn("metadata", enriched.sheetnames)
            headers = [cell.value for cell in enriched["community_tests"][1]]
            self.assertIn("internal comments", headers)
        finally:
            enriched.close()

    def test_missing_skip_reason_uses_contract_fallback(self) -> None:
        workbook = openpyxl.load_workbook(self.path)
        workbook["community_tests"]["F3"] = ""
        workbook.save(self.path)
        workbook.close()
        snapshot = WorkbookLoader().load(self.path, VERSION)
        semantic = build_deterministic_report(snapshot)
        semantic.skip_groups[0].redmine_urls = ["https://redmine.mellanox.com/issues/999"]
        path = ArtifactWriter().write_skips(snapshot, semantic)
        self.assertEqual(
            FALLBACK_SKIP_REASON,
            json.loads(path.read_text(encoding="utf-8"))[
                "https://redmine.mellanox.com/issues/999"
            ],
        )

    def test_skip_rows_fold_by_file_and_reason(self) -> None:
        workbook = openpyxl.load_workbook(self.path)
        sheet = workbook["community_tests"]
        for key, name, reason in (
            ("k4", "tests/b.py::test_other", "Internal ticket"),
            ("k5", "tests/b.py::test_param[a]", "Other condition"),
            ("k6", "tests/b.py::test_param[b]", "Internal ticket"),
        ):
            sheet.append(
                [
                    key,
                    key,
                    "tb2",
                    name,
                    "skipped",
                    reason,
                    "t1",
                    "dut",
                    "spc",
                    "x86",
                    "sku2",
                    VERSION,
                    name,
                    "",
                ]
            )
        workbook.save(self.path)
        workbook.close()

        semantic = build_deterministic_report(
            WorkbookLoader().load(self.path, VERSION)
        )
        self.assertEqual(1, len(semantic.skip_groups))
        self.assertEqual("tests/b.py::*", semantic.skip_groups[0].test_display)
        self.assertEqual(4, len(semantic.skip_groups[0].member_ids))
        self.assertEqual(
            "Internal ticket; Other condition",
            semantic.skip_groups[0].internal_comments,
        )

    def test_lifecycle_duplicates_share_comment_without_formula_injection(self) -> None:
        workbook = openpyxl.load_workbook(self.path)
        sheet = workbook["community_tests"]
        sheet.append([cell.value for cell in sheet[2]])
        workbook.save(self.path)
        workbook.close()

        snapshot = WorkbookLoader().load(self.path, VERSION)
        failure = snapshot.failures[0]
        self.assertEqual(2, len(snapshot.record_excel_rows[failure.record_id]))
        snapshot.row_comments[failure.record_id] = "=unsafe"
        semantic = build_deterministic_report(snapshot)
        writer = ArtifactWriter()
        attachment = writer.build_attachment(snapshot, semantic)
        self.addCleanup(writer.cleanup_attachment, attachment)
        enriched = openpyxl.load_workbook(attachment, data_only=False)
        try:
            sheet = enriched["community_tests"]
            comment_column = [cell.value for cell in sheet[1]].index("internal comments") + 1
            self.assertEqual("'=unsafe", sheet.cell(2, comment_column).value)
            self.assertEqual("'=unsafe", sheet.cell(5, comment_column).value)
        finally:
            enriched.close()


if __name__ == "__main__":
    unittest.main()
