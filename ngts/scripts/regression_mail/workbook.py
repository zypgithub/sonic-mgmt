"""Read-only Excel ingestion for collected SONiC results."""

from __future__ import annotations

import json
import os
import shutil
import tempfile
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, DefaultDict, Dict, Iterable, List, Mapping, Optional, Sequence, Tuple

from ngts.scripts.regression_mail.models import ResultRow, SemanticReport, WorkbookSnapshot
from ngts.scripts.regression_mail.normalization import normalize_text, normalize_version, stable_record_id


REQUIRED_HEADERS: Sequence[str] = (
    "session_id",
    "mars_key_id",
    "testbed",
    "test name",
    "result",
    "message",
    "topology",
    "host",
    "asic",
    "platform",
    "hwsku",
    "os_version",
    "sanitized_testname",
)
SUPPORTED_RESULTS = {"pass", "fail", "skipped"}
_RESULT_PRIORITY = {"pass": 0, "skipped": 1, "fail": 2}
FALLBACK_SKIP_REASON = "skipped by internal ticket，WIP。"


class WorkbookLoader:
    """Load the unique result sheet without mutating the source workbook."""

    def load(self, path: Path, version: str) -> WorkbookSnapshot:
        try:
            import openpyxl
        except ImportError as error:
            raise RuntimeError(
                "openpyxl is required; install ngts/scripts/regression_mail/requirements.txt"
            ) from error

        source = path.resolve()
        if source.suffix.lower() != ".xlsx":
            raise ValueError("Excel input must use the .xlsx extension")
        if not source.is_file():
            raise FileNotFoundError("Excel input does not exist: {}".format(source))

        workbook = openpyxl.load_workbook(source, read_only=True, data_only=False)
        try:
            sheet_name, header_row, headers = self._find_result_sheet(workbook)
            selected_rows, record_excel_rows = self._read_selected_rows(
                workbook[sheet_name],
                header_row,
                headers,
                normalize_version(version),
            )
        finally:
            workbook.close()

        if not selected_rows:
            raise ValueError("no Excel rows match version {!r}".format(version))

        counts = Counter(row.result for row in selected_rows)
        hardware_pairs = sorted(
            {(row.hwsku, row.topology) for row in selected_rows if row.hwsku or row.topology}
        )
        return WorkbookSnapshot(
            source_path=source,
            sheet_name=sheet_name,
            header_row=header_row,
            headers=tuple(headers),
            selected_rows=selected_rows,
            result_counts={name: counts.get(name, 0) for name in sorted(SUPPORTED_RESULTS)},
            hardware_pairs=[tuple(pair) for pair in hardware_pairs],
            record_excel_rows=record_excel_rows,
        )

    @staticmethod
    def _find_result_sheet(workbook: Any) -> Tuple[str, int, List[str]]:
        candidates: List[Tuple[str, int, List[str]]] = []
        required = set(REQUIRED_HEADERS)
        for worksheet in workbook.worksheets:
            for row_number, row in enumerate(
                worksheet.iter_rows(min_row=1, max_row=min(10, worksheet.max_row), values_only=True),
                start=1,
            ):
                headers = [str(value).strip() if value is not None else "" for value in row]
                if required.issubset(set(headers)):
                    if len([header for header in headers if header]) != len(
                        set(header for header in headers if header)
                    ):
                        raise ValueError(
                            "result sheet {!r} has duplicate column names".format(worksheet.title)
                        )
                    candidates.append((worksheet.title, row_number, headers))
                    break

        if not candidates:
            raise ValueError(
                "no worksheet contains all required columns: {}".format(", ".join(REQUIRED_HEADERS))
            )
        if len(candidates) != 1:
            raise ValueError(
                "multiple worksheets contain the required result columns: {}".format(
                    ", ".join(candidate[0] for candidate in candidates)
                )
            )
        return candidates[0]

    @staticmethod
    def _read_selected_rows(
        worksheet: Any,
        header_row: int,
        headers: Sequence[str],
        wanted_version: str,
    ) -> Tuple[List[ResultRow], Dict[str, List[int]]]:
        header_index = {header: index for index, header in enumerate(headers) if header}
        deduplicated: Dict[str, ResultRow] = {}
        record_excel_rows: Dict[str, List[int]] = {}
        for excel_row, cells in enumerate(
            worksheet.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1,
        ):
            values = {
                header: cells[index] if index < len(cells) else None
                for header, index in header_index.items()
            }
            row_version = normalize_version(values.get("os_version"))
            if row_version != wanted_version:
                continue

            result = normalize_text(values.get("result")).lower()
            if result not in SUPPORTED_RESULTS:
                raise ValueError(
                    "unsupported result {!r} at {}!{}".format(
                        values.get("result"),
                        worksheet.title,
                        excel_row,
                    )
                )

            test_name = normalize_text(values.get("test name"))
            session_id = normalize_text(values.get("session_id"))
            key_id = normalize_text(values.get("mars_key_id"))
            record_id = stable_record_id(row_version, session_id, key_id, test_name)
            record_excel_rows.setdefault(record_id, []).append(excel_row)
            row = ResultRow(
                record_id=record_id,
                excel_row=excel_row,
                session_id=session_id,
                mars_key_id=key_id,
                testbed=normalize_text(values.get("testbed")),
                test_name=test_name,
                sanitized_testname=normalize_text(values.get("sanitized_testname")),
                result=result,
                message=normalize_text(values.get("message")),
                topology=normalize_text(values.get("topology")),
                host=normalize_text(values.get("host")),
                asic=normalize_text(values.get("asic")),
                platform=normalize_text(values.get("platform")),
                hwsku=normalize_text(values.get("hwsku")),
                os_version=row_version,
                values=values,
            )
            existing = deduplicated.get(record_id)
            if existing is None or _RESULT_PRIORITY[row.result] > _RESULT_PRIORITY[existing.result]:
                deduplicated[record_id] = row

        return (
            sorted(deduplicated.values(), key=lambda row: row.excel_row),
            record_excel_rows,
        )


def iter_join_names(row: ResultRow) -> Iterable[str]:
    """Expose the exact and sanitized join names in deterministic order."""

    return row.join_names


class ArtifactWriter:
    """Create skip mappings and the enriched workbook copy."""

    def write_skips(
        self,
        workbook: WorkbookSnapshot,
        semantic: SemanticReport,
    ) -> Path:
        rows_by_id = {row.record_id: row for row in workbook.skipped}
        reasons: DefaultDict[str, List[str]] = defaultdict(list)
        for group in semantic.skip_groups:
            excel_reasons = sorted(
                {
                    normalize_text(rows_by_id[member].message)
                    for member in group.member_ids
                    if member in rows_by_id and normalize_text(rows_by_id[member].message)
                }
            )
            value = "; ".join(excel_reasons) if excel_reasons else FALLBACK_SKIP_REASON
            for url in group.redmine_urls:
                canonical = url.lower()
                if value not in reasons[canonical]:
                    reasons[canonical].append(value)

        payload = {
            url: "; ".join(sorted(set(values)))
            for url, values in sorted(reasons.items())
        }
        destination = workbook.source_path.parent / "skips.json"
        _atomic_json(destination, payload)
        return destination

    def build_attachment(
        self,
        workbook: WorkbookSnapshot,
        semantic: SemanticReport,
    ) -> Path:
        try:
            import openpyxl
        except ImportError as error:
            raise RuntimeError(
                "openpyxl is required; install ngts/scripts/regression_mail/requirements.txt"
            ) from error

        temp_dir = Path(tempfile.mkdtemp(prefix="sonic-regression-mail-"))
        destination = temp_dir / (
            workbook.source_path.stem + "_with_internal_comments.xlsx"
        )
        shutil.copy2(workbook.source_path, destination)
        comments = _comments_by_record(semantic)
        loaded = openpyxl.load_workbook(destination, data_only=False)
        try:
            worksheet = loaded[workbook.sheet_name]
            comment_column = None
            for column in range(1, worksheet.max_column + 1):
                if worksheet.cell(workbook.header_row, column).value == "internal comments":
                    comment_column = column
                    break
            if comment_column is None:
                comment_column = worksheet.max_column + 1
                worksheet.cell(workbook.header_row, comment_column).value = "internal comments"

            for row in workbook.selected_rows:
                value = comments.get(row.record_id, "")
                for excel_row in workbook.record_excel_rows.get(row.record_id, [row.excel_row]):
                    worksheet.cell(excel_row, comment_column).value = _safe_excel_text(value)
            loaded.save(destination)
        finally:
            loaded.close()
        return destination

    @staticmethod
    def cleanup_attachment(path: Optional[Path]) -> None:
        if path:
            shutil.rmtree(path.parent, ignore_errors=True)


def _comments_by_record(semantic: SemanticReport) -> Dict[str, str]:
    comments: Dict[str, str] = {}
    for group in list(semantic.failure_groups) + list(semantic.skip_groups):
        for member in group.member_ids:
            comments[member] = group.internal_comments
    return comments


def _safe_excel_text(value: object) -> str:
    text = str(value or "")
    return "'" + text if text.startswith(("=", "+", "-", "@")) else text


def _atomic_json(path: Path, payload: Mapping[str, str]) -> None:
    handle = tempfile.NamedTemporaryFile(
        mode="w",
        encoding="utf-8",
        prefix=".{}.".format(path.name),
        suffix=".tmp",
        dir=str(path.parent),
        delete=False,
    )
    temp_path = Path(handle.name)
    try:
        json.dump(payload, handle, indent=2, ensure_ascii=False, sort_keys=True)
        handle.write("\n")
        handle.flush()
        os.fsync(handle.fileno())
        handle.close()
        os.replace(temp_path, path)
    except Exception:
        handle.close()
        try:
            temp_path.unlink()
        except FileNotFoundError:
            pass
        raise
